#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import random
import socket
import time
import os
from urllib import request
from bs4 import BeautifulSoup
from openpyxl import Workbook, utils, load_workbook

# 声明变量
uaList = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
          {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
          {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

rd = random.randint(0, len(uaList) - 1)
ua = uaList[rd]


class Movie(object):
    def __init__(self, name, genre, year, download):
        self.name = name
        self.genre = genre
        self.year = year
        self.download = download

    def __str__(self):
        return str("%s, %s, %s" % (self.name, self.genre, self.year))


class MovieDownload(object):
    def __init__(self, file_name, resolution, size, magnet):
        self.file_name = file_name
        self.resolution = resolution
        self.size = size
        self.magnet = magnet


def escape_email(string):
    email = '[email'
    if email in string:
        return string[0:string.index(email) - 1]
    else:
        return string


def imdb_spider(imdb_url, rty_cnt=0):

    if rty_cnt > 10:
        return

    req = request.Request(imdb_url, headers=ua)

    ret_list = []
    try:
        with request.urlopen(req) as f:
            source_code = f.read().decode('UTF-8')

            soup = BeautifulSoup(source_code, 'html.parser')
            movie_list = soup.select('li > div')

            for movie_info in movie_list:
                # imdb_score = movie_info.select('.x-movie-mediumimg .imdb_index')[0].get_text()
                item = movie_info.select('.item-desc')[0]
                # name = item.p.a.string
                link = item.p.a['href']

                movie = movie_detail_spider(link)
                if movie:
                    ret_list.append(movie)
                    print('movie (%s) saved...' % movie)

            return ret_list

    except Exception as e:
        print('Error when requesting %s, Error:%s' % (imdb_url, e))
        rty_cnt += 1
        print('Retrying %s time...' % rty_cnt)
        return imdb_spider(imdb_url, rty_cnt)


def movie_detail_spider(link, rty_cnt=0):

    if rty_cnt > 10:
        return

    time.sleep(random.random() * 5)

    movie_req = request.Request(link, headers=ua)

    print('request movie %s' % link)
    try:
        with request.urlopen(movie_req) as f:

            # print('Status: %s %s' % (f.status, f.reason))

            source_code = f.read().decode('UTF-8')
            soup = BeautifulSoup(source_code, 'html.parser')

            row = soup.select('.row .row')[0]
            name = row.select('h2 a')[0].get_text()
            view_film = row.select('#viewfilm')[0]
            score = view_film.select('.badge')[0].get_text()
            imdb_score = view_film.select('.badge')[1].get_text()

            view_film_a = view_film.select('a')

            actor = []
            for a in view_film_a:
                href = a['href']
                if 'director' in href:
                    director = a.get_text()
                elif 'actor' in href:
                    actor.append(a.get_text())
                elif 'type' in href:
                    genre = a.get_text()
                elif 'country' in href:
                    country = a.get_text()
                elif 'year' in href:
                    year = a.get_text()

            movie_detail_list = soup.select('#cili tr')
            movie_download_list = []
            for movie_detail in movie_detail_list:
                try:
                    file_name = escape_email(movie_detail.select('b')[0].get_text())
                    resolution = movie_detail['id']
                    size = movie_detail.select('.label-warning')[0].get_text()
                    magnet = movie_detail.select('.btn-primary')[0]['href']
                    movie_download = MovieDownload(file_name, resolution, size, magnet)
                    movie_download_list.append(movie_download)
                except KeyError:
                    continue
                except IndexError:
                    continue

            movie = Movie(name, genre, year, movie_download_list)
            movie.director = director
            movie.actor = '/'.join(actor)
            movie.country = country
            movie.score = score
            movie.imdb_score = imdb_score
            return movie

    except Exception as e:
        print('Error when requesting %s, Error:%s' % (link, e))
        rty_cnt += 1
        print('Retrying %s time...' % rty_cnt)
        return movie_detail_spider(link, rty_cnt)


def save_to_excel(movie_list):

    ws_header = ['名称', '导演', '主演', '类型', '地区', '上映时间', '打分', 'IMDB评分', '分辨率', '大小', '下载地址', '文件名']

    save_path = 'IMDB电影.xlsx'
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(ws_header)

    for movie in movie_list:
        movie_download = movie.download
        for i in range(len(movie_download)):
            if i == 0:
                ws.append([movie.name, movie.director, movie.actor, movie.genre, movie.country, movie.year, movie.score,
                           movie.imdb_score, movie_download[i].resolution,
                           movie_download[i].size, movie_download[i].magnet, movie_download[i].file_name])
            else:
                ws.append(['', '', '', '', '', '', '', '', movie_download[i].resolution,
                           movie_download[i].size, movie_download[i].magnet, movie_download[i].file_name])

        for col in range(1, len(ws_header) - 3):
            end_row = ws.max_row
            start_row = end_row - len(movie_download) + 1
            col = utils.get_column_letter(col)

            start = col + str(start_row)
            end = col + str(end_row)

            ws.merge_cells(start + ':' + end)

    wb.save(save_path)


def do_spider():
    timeout = 15
    socket.setdefaulttimeout(timeout)
    page_no = 1

    while True:
        imdb_url = 'http://gaoqing.fm/ajax.php?p=' + str(page_no) + '&sort=IMDb'
        try:
            imdb_list = imdb_spider(imdb_url)
            if imdb_list:
                save_to_excel(imdb_list)
                page_no += 1
            else:
                break
        except Exception as e:
            print('Error in do_spider %s' % e)
            continue
        finally:
            print('Complete...')


if __name__ == '__main__':
    # print(movie_detail_spider('http://gaoqing.fm/view/502a83b56733'))
    do_spider()
